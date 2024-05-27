import pandas as pd
from openpyxl import Workbook
import advSearch as adv
import ragsEval as evalRag
def evalute_scores(excel_file):

    # Read the Excel file
    df = pd.read_excel(excel_file,engine='openpyxl')

    # Extract values from the 'query' column
    print(df.columns)
    query_values = df['Query'].dropna().tolist()
    ground_truth=df['GroundTruth'].dropna().to_list()
    print(query_values)
    
    chunk_json=[]
    # Write headers
    # ws.cell(row=1, column=1, value="400chunks-8chunksToLLM(adtalem)")
    wb = Workbook()
    ws = wb.active
    headers = ['query','SA_RESPONSE']
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
    queries=[]
    answers=[]
    ground_truth_values=[]
    contexts=[]
    context_url=[]
    for query,truth in zip(query_values,ground_truth):
        # print(query)
        data=adv.get_Bot_Response(query,truth)
        queries.append(data['query'])
        answers.append(data['answer'])
        ground_truth_values.append(data['ground_truth'])
        contexts.append(data['context'])
        context_url.append(data['context_url'])



    df=evalRag.evaluate_data(queries,answers,ground_truth_values,contexts,context_url)
    df.to_excel('output.xlsx', index=False)
    # print(len(chunk_json))
    # print(chunk_json)
    

    # Save the workbook
    # wb.save("AddtionalPrompt400To800.xlsx")
if __name__=="__main__":
    excel_file= r'/home/Ragul.Sivakumar/Documents/sa_input.xlsx'
    evalute_scores(excel_file)
